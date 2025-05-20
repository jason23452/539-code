import sys
import os
import time
import re
import math
import heapq
import itertools
import threading
from multiprocessing import Pool, cpu_count, freeze_support

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment

# -------------------------------------------------
# 全域變數
top_n = 200             # 要保留的最佳組合数
max_gap_limit = 1000000  # 最大相邻中奖期距阈值
lottery_masks = []      # 子进程初始化后存放掩码列表
chunk_size_for_combos = 100000
# -------------------------------------------------

def close_excel_workbook(file_path):
    if sys.platform.startswith('win'):
        try:
            import win32com.client
            excel = win32com.client.Dispatch('Excel.Application')
            fp = os.path.abspath(file_path).lower()
            for wb in list(excel.Workbooks):
                if wb.FullName.lower() == fp:
                    wb.Close(SaveChanges=True)
                    print(f'已关闭 Excel: {file_path}')
                    break
        except Exception as e:
            print('关闭 Excel 失败：', e)
    elif sys.platform.startswith('darwin'):
        try:
            import subprocess
            name = os.path.basename(os.path.abspath(file_path))
            script = f'tell application "Microsoft Excel" to close workbook "{name}" saving yes'
            subprocess.run(['osascript', '-e', script])
            print(f'已关闭 Excel: {file_path}')
        except Exception as e:
            print('关闭 Excel 失败：', e)
    else:
        print('不支持的操作系统，略过关闭。')

def reopen_excel_workbook(file_path):
    if sys.platform.startswith('win'):
        try:
            import win32com.client
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = True
            excel.Workbooks.Open(os.path.abspath(file_path))
            print(f'已重新开启 Excel: {file_path}')
        except Exception as e:
            print('重新开启 Excel 失败：', e)
    elif sys.platform.startswith('darwin'):
        try:
            import subprocess
            script = f'tell application "Microsoft Excel" to open POSIX file "{os.path.abspath(file_path)}"'
            subprocess.run(['osascript', '-e', script])
            print(f'已重新开启 Excel: {file_path}')
        except Exception as e:
            print('重新开启 Excel 失败：', e)
    else:
        print('不支持的操作系统，略过开启。')

def pad_data(data, total_rows, num_columns):
    while len(data) < total_rows:
        data.append([''] * num_columns)
    return data

def init_pool(l_masks):
    global lottery_masks
    lottery_masks = l_masks

def process_chunk(chunk_of_combos):
    '''多进程计算：维护 2星／3星／精确4星／精确5星 的 top_n 小堆'''
    heap2 = []; heap3 = []; heap4 = []; heap5 = []
    total = len(lottery_masks)
    for combo in chunk_of_combos:
        m = 0
        for n in combo:
            m |= 1 << (n - 1)
        cnt2 = cnt3 = cnt4 = cntE4 = cnt5 = cntE5 = 0
        last2 = last3 = lastE4 = last5 = lastE5 = -1
        for idx, lm in enumerate(lottery_masks, start=1):
            matches = (m & lm).bit_count()
            if matches >= 2:
                cnt2 += 1; last2 = idx
            if matches >= 3:
                cnt3 += 1; last3 = idx
            if matches >= 4:
                cnt4 += 1
                if matches == 4:
                    cntE4 += 1; lastE4 = idx
            if matches >= 5:
                cnt5 += 1; last5 = idx
                if matches == 5:
                    cntE5 += 1; lastE5 = idx
        diff2 = total - last2 if last2 != -1 else total
        diff3 = total - last3 if last3 != -1 else total
        diffE4 = total - lastE4 if lastE4 != -1 else total
        diff5 = total - last5 if last5 != -1 else total
        diffE5 = total - lastE5 if lastE5 != -1 else total
        item = (tuple(combo), cnt2, cnt3, cnt4, cntE4, cnt5, cntE5, diff2, diff3, diffE4, diff5, diffE5)
        # 2星堆
        key2 = (cnt2, cnt3, cnt4)
        if len(heap2) < top_n: heapq.heappush(heap2, (key2, item))
        elif key2 > heap2[0][0]: heapq.heapreplace(heap2, (key2, item))
        # 3星堆
        key3 = (cnt3, cnt4, cnt2)
        if len(heap3) < top_n: heapq.heappush(heap3, (key3, item))
        elif key3 > heap3[0][0]: heapq.heapreplace(heap3, (key3, item))
        # 4星堆 (精确4星)
        key4 = cntE4
        if len(heap4) < top_n: heapq.heappush(heap4, (key4, item))
        elif key4 > heap4[0][0]: heapq.heapreplace(heap4, (key4, item))
        # 5星堆 (精确5星)
        key5 = cntE5
        if len(heap5) < top_n: heapq.heappush(heap5, (key5, item))
        elif key5 > heap5[0][0]: heapq.heapreplace(heap5, (key5, item))
    return heap2, heap3, heap4, heap5

def compute_max_gap(combo, masks, threshold, exact=False):
    '''计算所有“差值”（相邻命中间隔），并返回最大差值'''
    m = 0
    for n in combo:
        m |= 1 << (n - 1)
    prev = 0
    diffs = []
    total = len(masks)
    for idx, lm in enumerate(masks, start=1):
        matches = (m & lm).bit_count()
        ok = (matches == threshold) if exact else (matches >= threshold)
        if ok:
            diffs.append(idx - prev)
            prev = idx
    diffs.append(total + 1 - prev)
    return max(diffs) if diffs else total

def main(sheet_range, combo_size, file_path):
    freeze_support()
    print(f'文件：{file_path}，范围：{sheet_range}，组合大小：{combo_size}')
    print(f'top_n={top_n}，max_gap_limit={max_gap_limit}')

    close_excel_workbook(file_path)
    time.sleep(0.2)

    wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
    ws = wb[wb.sheetnames[0]]
    rng = sheet_range.split('!',1)[-1].replace('$','')
    try:
        sc, ec = rng.split(':')
    except:
        print('范围格式错误'); sys.exit(1)
    sm = re.match(r'([A-Za-z]+)(\d+)?', sc)
    em = re.match(r'([A-Za-z]+)(\d+)?', ec)
    if not sm or not em:
        print('解析范围失败'); sys.exit(1)
    sr = int(sm.group(2) or 1)
    er = int(em.group(2) or ws.max_row)
    c1 = column_index_from_string(sm.group(1))
    c2 = column_index_from_string(em.group(1))
    it = ws.iter_rows(min_row=sr, max_row=er, min_col=c1, max_col=c2, values_only=True)
    try:
        headers = next(it)
    except StopIteration:
        print('无数据'); sys.exit(1)
    rows = list(it)
    wb.close()

    import pandas as pd
    df = pd.DataFrame(rows, columns=headers).dropna()
    draws = df.iloc[:, :combo_size].astype(int).values.tolist()

    masks = []
    for nums in draws:
        mm = 0
        for v in nums:
            if 1 <= v <= 39:
                mm |= 1 << (v - 1)
        masks.append(mm)

    from math import comb
    total = comb(39, combo_size)
    print(f'总组合数：{total}')
    all_combos = itertools.combinations(range(1,40), combo_size)

    def chunker(it, size):
        buf = []
        for x in it:
            buf.append(x)
            if len(buf) == size:
                yield buf
                buf = []
        if buf:
            yield buf

    t0 = time.time()
    with Pool(cpu_count(), initializer=init_pool, initargs=(masks,)) as pool:
        partials = pool.map(process_chunk, chunker(all_combos, chunk_size_for_combos))
    print(f'分布式计算耗时：{time.time()-t0:.2f}s')

    # ===== 先根据最大差值过滤，再取 top_n =====
    # 2星
    all2 = {}
    for heap2, _, _, _ in partials:
        for _, item in heap2:
            combo, cnt2, cnt3, cnt4, cntE4, cnt5, cntE5, diff2, diff3, diffE4, diff5, diffE5 = item
            gap2 = compute_max_gap(combo, masks, threshold=2)
            if gap2 <= max_gap_limit:
                all2[tuple(combo)] = (combo, cnt2, cnt3, cnt4, cnt5, diff2, gap2)
    sorted2 = sorted(all2.values(), key=lambda x: (x[1], x[2], x[3]), reverse=True)[:top_n]

    # 3星
    all3 = {}
    for _, heap3, _, _ in partials:
        for _, item in heap3:
            combo, cnt2, cnt3, cnt4, cntE4, cnt5, cntE5, diff2, diff3, diffE4, diff5, diffE5 = item
            gap3 = compute_max_gap(combo, masks, threshold=3)
            if gap3 <= max_gap_limit:
                all3[tuple(combo)] = (combo, cnt3, cnt4, cnt5, diff3, gap3)
    sorted3 = sorted(all3.values(), key=lambda x: (x[1], x[2]), reverse=True)[:top_n]

    # 4星（精确4星）
    all4 = {}
    for _, _, heap4, _ in partials:
        for _, item in heap4:
            combo, cnt2, cnt3, cnt4, cntE4, cnt5, cntE5, diff2, diff3, diffE4, diff5, diffE5 = item
            gap4 = compute_max_gap(combo, masks, threshold=4, exact=True)
            if gap4 <= max_gap_limit:
                all4[tuple(combo)] = (combo, cntE4, cnt5, diffE4, gap4)
    sorted4 = sorted(all4.values(), key=lambda x: x[1], reverse=True)[:top_n]

    # 5星（精确5星）
    all5 = {}
    for _, _, _, heap5 in partials:
        for _, item in heap5:
            combo, cnt2, cnt3, cnt4, cntE4, cnt5, cntE5, diff2, diff3, diffE4, diff5, diffE5 = item
            gap5 = compute_max_gap(combo, masks, threshold=5, exact=True)
            if gap5 <= max_gap_limit:
                all5[tuple(combo)] = (combo, cntE5, diffE5, gap5)
    sorted5 = sorted(all5.values(), key=lambda x: x[1], reverse=True)[:top_n]

    # 构造写入数据，同时保留“未开”（diff）和新增“最大差值”（gap）
    data2 = [ list(combo) + [cnt2, cnt3, cnt4, cnt5, diff2, gap2] for combo, cnt2, cnt3, cnt4, cnt5, diff2, gap2 in sorted2 ]
    data3 = [ list(combo) + [cnt3, cnt4, cnt5, diff3, gap3] for combo, cnt3, cnt4, cnt5, diff3, gap3 in sorted3 ]
    data4 = [ list(combo) + [cntE4, cnt5, diffE4, gap4] for combo, cntE4, cnt5, diffE4, gap4 in sorted4 ]
    data5 = [ list(combo) + [cntE5, diffE5, gap5] for combo, cntE5, diffE5, gap5 in sorted5 ]

    # 补足至 top_n 行
    data2 = pad_data(data2, top_n, combo_size + 6)
    data3 = pad_data(data3, top_n, combo_size + 5)
    data4 = pad_data(data4, top_n, combo_size + 4)
    data5 = pad_data(data5, top_n, combo_size + 3)

    # 写回 Excel
    wb2 = openpyxl.load_workbook(file_path, keep_vba=True)
    if '獲獎排列' in wb2.sheetnames:
        del wb2['獲獎排列']
    ws_out = wb2.create_sheet('獲獎排列', index=1)

    col = 1
    # 2星表头
    hdr2 = [f'號碼{i}' for i in range(1, combo_size+1)] + ['2星','3星','4星','5星','未開','最大差值']
    for j, h in enumerate(hdr2, start=col): ws_out.cell(1, j, h).alignment = Alignment('center','center')
    for i, row in enumerate(data2, start=2):
        for j, v in enumerate(row, start=col): ws_out.cell(i, j, v).alignment = Alignment('center','center')
    
    # 3星表头
    col += len(hdr2) + 1
    hdr3 = [f'號碼{i}' for i in range(1, combo_size+1)] + ['3星','4星','5星','未開','最大差值']
    for j, h in enumerate(hdr3, start=col): ws_out.cell(1, j, h).alignment = Alignment('center','center')
    for i, row in enumerate(data3, start=2):
        for j, v in enumerate(row, start=col): ws_out.cell(i, j, v).alignment = Alignment('center','center')

    # 4星表头
    col += len(hdr3) + 1
    hdr4 = [f'號碼{i}' for i in range(1, combo_size+1)] + ['4星','5星','未開','最大差值']
    for j, h in enumerate(hdr4, start=col): ws_out.cell(1, j, h).alignment = Alignment('center','center')
    for i, row in enumerate(data4, start=2):
        for j, v in enumerate(row, start=col): ws_out.cell(i, j, v).alignment = Alignment('center','center')

    # 5星表头
    col += len(hdr4) + 1
    hdr5 = [f'號碼{i}' for i in range(1, combo_size+1)] + ['5星','未開','最大差值']
    for j, h in enumerate(hdr5, start=col): ws_out.cell(1, j, h).alignment = Alignment('center','center')
    for i, row in enumerate(data5, start=2):
        for j, v in enumerate(row, start=col): ws_out.cell(i, j, v).alignment = Alignment('center','center')

    wb2.save(file_path)
    print('已寫入「獲獎排列」並保存完成。')
    reopen_excel_workbook(file_path)

if __name__ == '__main__':
    freeze_support()
    if len(sys.argv) < 4:
        print('用法：<SheetRange> <combo_size> <excel_path> [top_n] [max_gap_limit]')
        sys.exit(1)

    sheet_range = sys.argv[1]
    combo_size = int(sys.argv[2])
    excel_path = sys.argv[3]

    if len(sys.argv) >= 5:
        try:
            top_n = int(sys.argv[4])
        except ValueError:
            print('第4个参数 top_n 必须是整数'); sys.exit(1)
    if len(sys.argv) >= 6:
        try:
            max_gap_limit = int(sys.argv[5])
        except ValueError:
            print('第5个参数 max_gap_limit 必须是整数'); sys.exit(1)

    import tkinter as tk
    from tkinter.scrolledtext import ScrolledText
    root = tk.Tk(); root.title('计算终端')
    ta = ScrolledText(root, wrap='word', font=('Consolas',10)); ta.pack(expand=True, fill='both')
    class R:
        def __init__(self, w): self.w = w
        def write(self, s): self.w.after(0, self.w.insert, tk.END, s); self.w.after(0, self.w.see, tk.END)
        def flush(self): pass
    sys.stdout = R(ta); sys.stderr = R(ta)
    threading.Thread(target=lambda: main(sheet_range, combo_size, excel_path), daemon=True).start()
    tk.Button(root, text='退出', command=root.destroy).pack(pady=5)
    root.mainloop()
