import sys
import os
import re
import subprocess
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import messagebox

def close_excel_workbook(file_path):
    """嘗試關閉已開啟的 Excel 活頁簿，確保檔案可供寫入"""
    if sys.platform.startswith('win'):
        try:
            import win32com.client
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = False
            fp = os.path.abspath(file_path).lower()
            for wb in list(excel.Workbooks):
                if wb.FullName.lower() == fp:
                    wb.Close(SaveChanges=True)
            excel.Quit()
        except Exception:
            pass
    elif sys.platform.startswith('darwin'):
        name = os.path.basename(os.path.abspath(file_path))
        script = f'tell application "Microsoft Excel" to close workbook "{name}" saving yes'
        subprocess.run(['osascript', '-e', script], check=False)

def reopen_excel_workbook(file_path):
    """略與關閉對應，根據系統實現重新開啟"""
    if sys.platform.startswith('win'):
        try:
            import win32com.client
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = True
            excel.Workbooks.Open(os.path.abspath(file_path))
        except Exception:
            pass
    elif sys.platform.startswith('darwin'):
        script = f'tell application "Microsoft Excel" to open POSIX file "{os.path.abspath(file_path)}"'
        subprocess.run(['osascript', '-e', script], check=False)

def detect_combo_size(ws):
    size = 0
    while ws.cell(1, size+1).value == f"號碼{size+1}":
        size += 1
    return size

def read_draws(ws, col_range):
    m = re.match(r'^([A-Za-z]+)(\d*):([A-Za-z]+)(\d*)$', col_range)
    if not m:
        raise ValueError(f"不支援的欄位範圍格式：{col_range}")
    c1, r1_str, c2, r2_str = m.groups()
    c1_i = column_index_from_string(c1)
    c2_i = column_index_from_string(c2)
    r1 = int(r1_str) if r1_str else 2
    r2 = int(r2_str) if r2_str else ws.max_row
    draws = []
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1_i, max_col=c2_i, values_only=True):
        if any(row):
            draws.append([int(v) for v in row if v is not None])
    return draws

def read_section_combos(ws, start_col, combo_size):
    combos = []
    end_col = start_col + combo_size - 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=start_col, max_col=end_col,
                            values_only=True):
        if all(v is not None for v in row):
            combos.append(tuple(int(v) for v in row))
    return combos

def count_hits(draws, combo, threshold, exact=False):
    cnt = 0
    for d in draws:
        hits = len(set(combo) & set(d))
        if (exact and hits == threshold) or (not exact and hits >= threshold):
            cnt += 1
    return cnt

def write_section(ws, combo_size, draws, combos, thresholds, start_col):
    headers = [f"號碼{i}" for i in range(1, combo_size+1)] + [name for name, *_ in thresholds]
    for j, h in enumerate(headers, start=start_col):
        ws.cell(1, j, h).alignment = Alignment(horizontal="center")
    for i, combo in enumerate(combos, start=2):
        for k, num in enumerate(combo, start=start_col):
            ws.cell(i, k, num).alignment = Alignment(horizontal="center")
        for idx, (_, thr, exact) in enumerate(thresholds, start=1):
            col = start_col + combo_size + idx - 1
            ws.cell(i, col, count_hits(draws, combo, thr, exact)).alignment = Alignment(horizontal="center")

def main(path, draws_sheet, col_range, prize_sheet):
    if not os.path.exists(path):
        messagebox.showerror("錯誤", "找不到檔案")
        return
    close_excel_workbook(path)
    wb = openpyxl.load_workbook(path, keep_vba=True)
    ws_draws = wb[draws_sheet]
    ws_prize = wb[prize_sheet]

    M = detect_combo_size(ws_prize)
    draws = read_draws(ws_draws, col_range)

    starts = [col for col in range(1, ws_prize.max_column+1)
              if ws_prize.cell(1, col).value == '號碼1']
    if len(starts) < 4:
        messagebox.showerror("錯誤", f"偵測到 {len(starts)} 個號碼段，無法回測四段。")
        return
    starts.sort()
    config = [
        [("2星", 2, False), ("3星", 3, False), ("4星", 4, False), ("5星", 5, True)],
        [("3星", 3, False), ("4星", 4, False), ("5星", 5, True)],
        [("4星", 4, False), ("5星", 5, True)],
        [("5星", 5, True)]
    ]

    if "回測結果" in wb.sheetnames:
        del wb["回測結果"]
    ws_result = wb.create_sheet("回測結果", index=len(wb.sheetnames))

    for start_col, thresholds in zip(starts[:4], config):
        combos = read_section_combos(ws_prize, start_col, M)
        write_section(ws_result, M, draws, combos, thresholds, start_col)

    try:
        wb.save(path)
    except PermissionError:
        close_excel_workbook(path)
        wb.save(path)
    reopen_excel_workbook(path)

    # 顯示完成訊息
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("回測完成", "所有回測區段已成功完成！")
    root.destroy()

if __name__ == '__main__':
    # GUI 入口
    root = tk.Tk()
    root.withdraw()
    if len(sys.argv) != 5:
        messagebox.showinfo("使用說明", "用法：python backtest_sections.py <檔案> <原始表> <範圍> <排列表>")
        sys.exit(1)
    _, path, draws_sheet, col_range, prize_sheet = sys.argv
    main(path, draws_sheet, col_range, prize_sheet)